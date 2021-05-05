# https://www.kaggle.com/promptcloud/careerbuilder-job-listing-2020
# Careerbuilder job listing
import argparse
import csv
import sys

import openpyxl as openpyxl


class NoCreativeName:
    class Entry:
        def __init__(self, no, date, country, last_update, more_time, confirmed, deaths, recovered):
            self.no = no
            self.date = date
            self.country = country
            self.last_update = last_update + more_time
            self.confirmed = confirmed
            self.deaths = int(deaths.split(".")[0])
            self.recovered = int(recovered.split(".")[0])

        def __repr__(self):
            return f'ID: {self.no}, Date: {self.date}, Place: {self.country}, Last update: {self.last_update}, Confirmed: {self.confirmed}, Deaths: {self.deaths}, Recovered: {self.recovered}'

    def __init__(self, name):
        self.data = []
        self.read_file(name)

        # self.deaths_by_date()
        # self.recovered_total()
        # self.countries_checked()

    def read_file(self, name):
        try:
            with open(name, newline='') as csvfile:
                spamreader = csv.reader(csvfile, delimiter=' ', quotechar='|')
                count = 0
                for row in spamreader:
                    count += 1
                    if count == 1:
                        continue
                    if count > 50_000:
                        break
                    data = []
                    for elem in row:
                        for word in elem.split(","):
                            data.append(word)
                    while len(data) > 8:
                        data.pop(2)
                    self.data.append(self.Entry(*data))
        except Exception as e:
            print(e)

        # self.data.pop(0)
        return self.data

    def recovered_total(self):
        sum = 0
        for entry in self.data:
            sum += entry.recovered
        return sum

    def deaths_by_date(self):
        deaths = dict()
        occurences = dict()

        for entry in self.data:
            if entry.date not in self.data:
                deaths[entry.date] = 0
            deaths[entry.date] += entry.deaths
        return deaths

    def countries_checked(self):
        countries = set()
        for entry in self.data:
            countries.add(entry.country)
        return len(countries)

    def save_to_excel(self, name):
        f_title = openpyxl.styles.Font(color="FF0000", italic=True, bold=True)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Covid global statistics"
        all_recovered = sheet.cell(row=1, column=3)
        all_recovered.value = "All recovered people"
        all_recovered.font = f_title
        countries_number = sheet.cell(row=1, column=4)
        countries_number.value = "Total Number of Cities"
        countries_number.font = f_title
        cities_number_value = sheet.cell(row=2, column=4)
        cities_number_value.value = self.countries_checked()
        sheet.merge_cells('A1:B1')
        deaths_by_date = sheet.cell(row=1, column=1)
        deaths_by_date.value = "Deaths per day"
        deaths_by_date.font = f_title
        deaths_by_date = sheet.cell(row=2, column=1)
        deaths_by_date.value = "Day"
        deaths_by_date = sheet.cell(row=2, column=2)
        deaths_by_date.value = "Deaths"
        counter = 3
        for date, deaths in self.deaths_by_date().items():
            sheet.cell(row=counter, column=1, value=date)
            sheet.cell(row=counter, column=2, value=deaths)
            counter += 1
        workbook.save(name)

def run():
    name = "a.csv"
    NoCreativeName(name)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="The idea of the application is to analyse data from CSV File,"
                                                 " perform few operations and have the option"
                                                 " to store the result of such operations to an Exel File")
    parser.add_argument('csv_file', type=str, help="The name of the CSV File")
    parser.add_argument('-o', help="To save results of all operations in the Exel File")
    args = parser.parse_args()
    file_name = args.csv_file
    if file_name[-4:] != '.csv':
        print("The file has a wrong format")
        sys.exit(0)

    no_name = NoCreativeName(file_name)
    # if 'o' in args and args.o is not None:
    #     name_xlsx_file = args.o
    #     no_name.save_to_excel(name_xlsx_file)